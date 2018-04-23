import MySQLdb
from openpyxl import load_workbook

db=MySQLdb.connect('localhost','root','balu','project')
cursor=db.cursor()

wb = load_workbook(filename="pytest.xlsx")
ws = wb['Mysheet']
for row in ws.iter_rows(min_row=2, min_col=1,max_row=1471,max_col=9):
    age=str(row[0].value)
    btravel=row[1].value
    dept=row[2].value
    gender=row[3].value
    jobrole=row[4].value
    hrate=str(row[5].value)
    drate=str(row[6].value)
    mincome=str(row[7].value)
    mrate=str(row[8].value)
    query="insert into big_Data values("+age+",'"+btravel+"','"+dept+"','"+gender+"','"+jobrole+"',"+hrate+","+drate+","+mincome+","+mrate+");"
    cursor.execute(query)

cursor.close()
db.commit()
db.close()

